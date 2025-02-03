import pandas as pd
import openpyxl
# from collections import deque

def load_workbook(file_path):
    return openpyxl.load_workbook(file_path, data_only=True)

def extract_columns(sheet):
    ISIN, BS, Quant, Date, Net_Amt, Unit_Price, symbols = [], [], [], [], [], [], []

    for i in range(2, 15):
        if i in {6, 7, 8, 10, 11, 12}:
            continue
        for row in sheet.iter_rows(min_col=i, max_col=i, min_row=3, values_only=True, max_row=9):
            for cell in row:
                if i == 2:
                    ISIN.append(cell)
                if i == 3:
                    symbols.append(cell)
                elif i == 4:
                    BS.append(cell)
                elif i == 5:
                    Quant.append(cell)
                elif i == 9:
                    Date.append(cell)
                elif i == 13:
                    Net_Amt.append(cell)
                elif i == 14:
                    Unit_Price.append(cell)
    
    return ISIN, symbols, BS, Quant, Date, Net_Amt, Unit_Price

def create_dataframe(ISIN, symbols, BS, Quant, Date, Net_Amt, Unit_Price):
    data = {
        'ISIN': ISIN,
        'symbols': symbols,
        'BS': BS,
        'Quant': Quant,
        'Date': Date,
        'Net_Amt': Net_Amt,
        'Unit_Price': Unit_Price
    }
    return pd.DataFrame(data)

def segregate_transactions(df):
    return df[df['BS'] == 'B'], df[df['BS'] == 'S']

def clear_previous_data(sheet):
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=3, max_col=8):
        for cell in row:
            cell.value = None 
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=10, max_col=15):
        for cell in row:
            cell.value = None

def write_to_sheet(sheet, df, columns, start_col):
    for col, col_name in enumerate(columns, start=start_col):
        for row_num, value in enumerate(df[col_name], start=4):
            sheet.cell(row=row_num, column=col, value=value)

def main():
    # Load the workbooks
    buy_sell_wb = load_workbook('Buy-sell.xlsx')
    holdings_wb = load_workbook('Holdings and valuation.xlsx')
    
    # Extract data from 'Buy-Sell Entry' sheet
    buy_sell_sheet = buy_sell_wb['Buy-Sell Entry']
    ISIN, symbols, BS, Quant, Date, Net_Amt, Unit_Price = extract_columns(buy_sell_sheet)
    
    # Create a DataFrame and segregate transactions
    q = create_dataframe(ISIN, symbols, BS, Quant, Date, Net_Amt, Unit_Price)
    Buy_q, Sell_q = segregate_transactions(q)
    
    # Prepare the 'Holdings, Valuations and P&L' sheet
    holdings_sheet = holdings_wb['Holdings, Valuations and P&L ']
    clear_previous_data(holdings_sheet)
    
    # Define the columns to write
    columns_to_write = ['ISIN', 'symbols', 'Quant', 'Date', 'Net_Amt', 'Unit_Price']
    
    # Write Buy and Sell transactions to the sheet
    write_to_sheet(holdings_sheet, Buy_q, columns_to_write, start_col=3)
    write_to_sheet(holdings_sheet, Sell_q, columns_to_write, start_col=10)
    
    # Save the workbook
    holdings_wb.save("Holdings and valuation.xlsx")

if __name__ == "__main__":
    main()
