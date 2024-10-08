#Importing reqd libraries
import pandas as pd
import openpyxl
import datetime as dt


def load_sheet_data(file_path, sheet_name, min_row, max_row):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        data = sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True)
        return wb, sheet, data
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


def process_trade_data(data, col_indices, columns, bs_column='BS', remove_columns=['BS', 'Net_Amt']):
    data_list = [[row[i] for i in col_indices] for row in data]
    df = pd.DataFrame(data_list, columns=columns)

    Buy_q = df[df[bs_column] == 'B']
    Sell_q = df[df[bs_column] == 'S']
    
    Buy_q = Buy_q.drop(columns=remove_columns, axis=1)
    Sell_q = Sell_q.drop(columns=remove_columns, axis=1)
    
    Buy_q['Date'] = pd.to_datetime(Buy_q['Date']).dt.date
    Sell_q['Date'] = pd.to_datetime(Sell_q['Date']).dt.date
    
    q_buy = [group.values.tolist() for _, group in Buy_q.groupby('ISIN')]
    q_sell = [group.values.tolist() for _, group in Sell_q.groupby('ISIN')]
    
    return df, q_buy, q_sell


def formating_queues(q_buy, q_sell):
    q_buy_formatted = []
    q_sell_formatted = []

    for group in q_buy:
        formatted_group = [[item if not isinstance(item, dt.date) else item.strftime('%d-%m-%Y') for item in row] for row in group]
        q_buy_formatted.append(formatted_group)

    for group in q_sell:
        formatted_group = [[item if not isinstance(item, dt.date) else item.strftime('%d-%m-%Y') for item in row] for row in group]
        q_sell_formatted.append(formatted_group)

    return q_buy_formatted, q_sell_formatted



def main_logic(q_sell_formatted, q_buy_formatted):
    residual_sell = []
    residual_buy = []

    while q_sell_formatted:
        for i in range(len(q_sell_formatted) - 1, -1, -1):  # Iterate in reverse to avoid index issues
            for j in range(len(q_buy_formatted) - 1, -1, -1):  # Iterate in reverse to avoid index issues
                # Check if lists are not empty before accessing elements
                if q_sell_formatted[i] and q_buy_formatted[j]:
                    if q_sell_formatted[i][0][0] == q_buy_formatted[j][0][0]:
                        if q_sell_formatted[i][0][2] == q_buy_formatted[j][0][2]:
                            residual_sell.append(q_sell_formatted[i].pop(0))
                            residual_buy.append(q_buy_formatted[j].pop(0))
                            break  # Exit the inner loop after popping elements
                        elif q_sell_formatted[i][0][2] < q_buy_formatted[j][0][2]:
                            q_buy_formatted[j][0][2] -= q_sell_formatted[i][0][2]
                            residual_buy.append([q_buy_formatted[j][0][0], q_buy_formatted[j][0][1], q_sell_formatted[i][0][2], q_buy_formatted[j][0][3], q_buy_formatted[j][0][4]])
                            residual_sell.append(q_sell_formatted[i].pop(0))
                            break  # Exit the inner loop after popping elements
                        else:
                            while q_sell_formatted[i] and q_sell_formatted[i][0][2] > q_buy_formatted[j][0][2]:  # Check if q_sell_formatted[i] is not empty
                                q_sell_formatted[i][0][2] -= q_buy_formatted[j][0][2]
                                residual_sell.append([q_sell_formatted[i][0][0], q_sell_formatted[i][0][1], q_buy_formatted[j][0][2], q_sell_formatted[i][0][3], q_sell_formatted[i][0][4]])
                                residual_buy.append(q_buy_formatted[j].pop(0))
                                if not q_buy_formatted[j]:  # Break if q_buy_formatted[j] becomes empty
                                    break
                            if q_buy_formatted[j]:  # Only proceed if there are still elements in q_buy_formatted[j]
                                if q_sell_formatted[i][0][2] == q_buy_formatted[j][0][2]:
                                    residual_sell.append(q_sell_formatted[i].pop(0))
                                    residual_buy.append(q_buy_formatted[j].pop(0))
                                    break
                                elif q_sell_formatted[i][0][2] < q_buy_formatted[j][0][2]:
                                    q_buy_formatted[j][0][2] -= q_sell_formatted[i][0][2]
                                    residual_buy.append([q_buy_formatted[j][0][0], q_buy_formatted[j][0][1], q_sell_formatted[i][0][2], q_buy_formatted[j][0][3], q_buy_formatted[j][0][4]])
                                    residual_sell.append(q_sell_formatted[i].pop(0))
                                    break
                else:
                    continue

        # Remove empty lists from q_sell_formatted and q_buy_formatted
        q_sell_formatted = [x for x in q_sell_formatted if x]
        q_buy_formatted = [x for x in q_buy_formatted if x]
    
    q_buy_final_2D = []
    for sublist1 in q_buy_formatted:
        for sublist2 in sublist1:
            q_buy_final_2D.append(sublist2)

    return residual_buy, residual_sell, q_buy_final_2D


def Amt(quant, unit):
   Net_Amt = quant*unit
   return Net_Amt


def pnl(sp, cp):
   a = sp-cp
   return a


def per_pnl(sp, cp):
   a = (sp-cp)/cp*100
   return a


def process_queues(residual_buy, residual_sell, q_buy_final_2D):

    df_q_buy_final = pd.DataFrame(q_buy_final_2D, columns=['ISIN', 'symbol','Quant', 'Date', 'unitprice'])
    df_q_buy_final['Date'] = pd.to_datetime(df_q_buy_final['Date'], format='%d-%m-%Y')
    df_residual_sell = pd.DataFrame(residual_sell, columns=['S.ISIN','S.symbol', 'S.Quant', 'S.Date', 'S.unitprice'])
    df_residual_sell['S.Date'] = pd.to_datetime(df_residual_sell['S.Date'], format='%d-%m-%Y')
    df_residual_buy = pd.DataFrame(residual_buy, columns=['B.ISIN','B.symbol', 'B.Quant', 'B.Date', 'B.unitprice'])
    df_residual_buy['B.Date'] = pd.to_datetime(df_residual_buy['B.Date'], format='%d-%m-%Y')


    #Adding Net_amt, p_l and per_p_l
    Net_Amt_Residual_buy = Amt(df_residual_buy['B.Quant'],df_residual_buy['B.unitprice'])
    df_residual_buy['B.NetAmt'] = Net_Amt_Residual_buy

    Net_Amt_Residual_Sell = Amt(df_residual_sell['S.Quant'],df_residual_sell['S.unitprice'])
    df_residual_sell['S.NetAmt'] = Net_Amt_Residual_Sell

    Net_Amt_final_buy = Amt(df_q_buy_final['Quant'],df_q_buy_final['unitprice'])
    df_q_buy_final['NetAmt'] = Net_Amt_final_buy

    pl = pnl(df_residual_sell['S.NetAmt'], df_residual_buy['B.NetAmt'])
    df_residual_sell['S.pl'] = pl

    per_pl = per_pnl(df_residual_sell['S.NetAmt'], df_residual_buy['B.NetAmt'])
    df_residual_sell['S.per_pl'] = per_pl

    #sorting
    df_q_buy_final.sort_values(by = 'Date', inplace = True, ascending=True)


    #Merging and sorting
    merged_q_buy_sell = pd.concat([df_residual_buy, df_residual_sell], axis = 1)
    merged_q_buy_sell.sort_values(by = 'S.Date', inplace = True, ascending=True)

    return df_q_buy_final, merged_q_buy_sell



def clear_columns(sheet, col_start1, col_end1, col_start2, col_end2, min_row = 4):
    for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row, min_col=col_start1, max_col=col_end1):
        for cell in row:
            cell.value = None 




def populate_sheet(wb, file_path, sheet, q_list, q_list1, df_q_buy_final, merged_q_buy_sell):
    col_mapping1 = {
        'ISIN': 17,
        'Quant': 19,
        'Date': 20,
        'unitprice': 21,
        'NetAmt': 22,
        'symbol': 18
    }

    col_mapping2 = {
        'S.ISIN': 34,
        'S.Quant': 36,
        'B.Date': 37,
        'B.unitprice': 38,
        'B.NetAmt': 39,
        'S.Date': 40,
        'S.unitprice': 41,
        'S.NetAmt': 42,
        'S.pl': 43,
        'S.per_pl': 44,
        'S.symbol': 35
    }

    for i in q_list:
        if i in col_mapping1:
            col = col_mapping1[i]
            for row_num, value in enumerate(df_q_buy_final[i], start=4):
                if i == 'Date' and isinstance(value, dt.date):
                    value = value.strftime('%d-%m-%Y')
                sheet.cell(row=row_num, column=col, value=value)
    
    for i in q_list1:
        if i in col_mapping2:
            col = col_mapping2[i]
            for row_num, value in enumerate(merged_q_buy_sell[i], start=4):
                if i in ['B.Date', 'S.Date'] and isinstance(value, dt.date):
                    value = value.strftime('%d-%m-%Y')
                sheet.cell(row=row_num, column=col, value=value)

    wb.save(file_path)