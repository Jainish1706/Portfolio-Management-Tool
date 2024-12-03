#Importing Libraries
import pandas as pd
import openpyxl
import pyxirr
from datetime import datetime, timedelta
import datetime as dt
import yfinance as yf




#This function only reads the sheet
def load_worksheet(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    return wb, sheet


#Function to read data from excel file
def extract_columns_to_dataframe(sheet, column_mapping, min_row, max_row, skip_columns=None):
    if skip_columns is None:
        skip_columns = []

    # Initialize columns lists
    column_data = {name: [] for name in column_mapping.values()}

    # Extract data from the sheet
    for i in range(1, max(column_mapping.keys()) + 1):
        if i in skip_columns:
            continue
        if i in column_mapping:
            for row in sheet.iter_rows(min_col=i, max_col=i, min_row=min_row, values_only=True, max_row=max_row):
                for cell in row:
                    column_name = column_mapping[i]
                    column_data[column_name].append(cell)

    # Convert to DataFrame
    df = pd.DataFrame(column_data)
    
    return df
    



#This function is segregating buy and sell queue
def buy_and_sell_q(df, bs_column='BS'):

    Buy_q = df[df[bs_column] == 'B']
    Sell_q = df[df[bs_column] == 'S']

    return Buy_q, Sell_q


#This segregates buy and sell queue with extra modifications to remove 
def process_trade_data(df, bs_column='BS', remove_columns=['BS', 'Net_Amt']):
    Buy_q = df[df[bs_column] == 'B']
    Sell_q = df[df[bs_column] == 'S']
    
    Buy_q = Buy_q.drop(columns=remove_columns, axis=1)
    Sell_q = Sell_q.drop(columns=remove_columns, axis=1)
    
    Buy_q['Date'] = pd.to_datetime(Buy_q['Date']).dt.date
    Sell_q['Date'] = pd.to_datetime(Sell_q['Date']).dt.date
    
    q_buy = [group.values.tolist() for _, group in Buy_q.groupby('ISIN')]
    q_sell = [group.values.tolist() for _, group in Sell_q.groupby('ISIN')]
    
    return q_buy, q_sell




#Function to change date from datetime object to string format
def formating_queues(q_buy, q_sell):
    q_buy_formatted = []
    q_sell_formatted = []

    for group in q_buy:
        formatted_group = [[item if not isinstance(item, dt.date) else item.strftime('%d-%m-%Y') for item in row] for row in group]
        q_buy_formatted.append(formatted_group)

    for group in q_sell:
        formatted_group = [[item if not isinstance(item, dt.date) else item.strftime('%d-%m-%Y') for item in row] for row in group]
        q_sell_formatted.append(formatted_group)

    return q_buy_formatted, q_sell_formatted





#This function contains the main logic of FIFO data structure to write in the excel file in proper manner
def process_sell_buy_orders(q_sell_formatted, q_buy_formatted):

    residual_sell = []
    residual_buy = []

    while q_sell_formatted:
        for i in range(len(q_sell_formatted) - 1, -1, -1):  # Iterate in reverse to avoid index issues
            for j in range(len(q_buy_formatted) - 1, -1, -1):  # Iterate in reverse to avoid index issues
                # Check if lists are not empty before accessing elements
                if q_sell_formatted[i] and q_buy_formatted[j]:
                    if q_sell_formatted[i][0][0] == q_buy_formatted[j][0][0]:
                        if q_sell_formatted[i][0][2] == q_buy_formatted[j][0][2]:
                            residual_sell.append(q_sell_formatted[i].pop(0))
                            residual_buy.append(q_buy_formatted[j].pop(0))
                            break  # Exit the inner loop after popping elements
                        elif q_sell_formatted[i][0][2] < q_buy_formatted[j][0][2]:
                            q_buy_formatted[j][0][2] -= q_sell_formatted[i][0][2]
                            residual_buy.append([q_buy_formatted[j][0][0], q_buy_formatted[j][0][1], q_sell_formatted[i][0][2], q_buy_formatted[j][0][3], q_buy_formatted[j][0][4]])
                            residual_sell.append(q_sell_formatted[i].pop(0))
                            break  # Exit the inner loop after popping elements
                        else:
                            while q_sell_formatted[i] and q_sell_formatted[i][0][2] > q_buy_formatted[j][0][2]:  # Check if q_sell_formatted[i] is not empty
                                q_sell_formatted[i][0][2] -= q_buy_formatted[j][0][2]
                                residual_sell.append([q_sell_formatted[i][0][0], q_sell_formatted[i][0][1], q_buy_formatted[j][0][2], q_sell_formatted[i][0][3], q_sell_formatted[i][0][4]])
                                residual_buy.append(q_buy_formatted[j].pop(0))
                                if not q_buy_formatted[j]:  # Break if q_buy_formatted[j] becomes empty
                                    break
                            if q_buy_formatted[j]:  # Only proceed if there are still elements in q_buy_formatted[j]
                                if q_sell_formatted[i][0][2] == q_buy_formatted[j][0][2]:
                                    residual_sell.append(q_sell_formatted[i].pop(0))
                                    residual_buy.append(q_buy_formatted[j].pop(0))
                                    break
                                elif q_sell_formatted[i][0][2] < q_buy_formatted[j][0][2]:
                                    q_buy_formatted[j][0][2] -= q_sell_formatted[i][0][2]
                                    residual_buy.append([q_buy_formatted[j][0][0], q_buy_formatted[j][0][1], q_sell_formatted[i][0][2], q_buy_formatted[j][0][3], q_buy_formatted[j][0][4]])
                                    residual_sell.append(q_sell_formatted[i].pop(0))
                                    break
                else:
                    continue

        # Remove empty lists from q_sell_formatted and q_buy_formatted
        q_sell_formatted = [x for x in q_sell_formatted if x]
        q_buy_formatted = [x for x in q_buy_formatted if x]

    # Flatten q_buy_formatted into 2D
    q_buy_final_2D = [sublist for sublist1 in q_buy_formatted for sublist in sublist1]

    return residual_sell, residual_buy, q_buy_final_2D





#This function merges residual_sell and residual_buy queues and sort it according to sell date. Also it sorts buy_final queue according to Date.
def process_queues(residual_buy, residual_sell, q_buy_final_2D):

    df_q_buy_final = pd.DataFrame(q_buy_final_2D, columns=['ISIN', 'symbol','Quant', 'Date', 'unitprice'])
    df_q_buy_final['Date'] = pd.to_datetime(df_q_buy_final['Date'], format='%d-%m-%Y')
    df_residual_sell = pd.DataFrame(residual_sell, columns=['S.ISIN','S.symbol', 'S.Quant', 'S.Date', 'S.unitprice'])
    df_residual_sell['S.Date'] = pd.to_datetime(df_residual_sell['S.Date'], format='%d-%m-%Y')
    df_residual_buy = pd.DataFrame(residual_buy, columns=['B.ISIN','B.symbol', 'B.Quant', 'B.Date', 'B.unitprice'])
    df_residual_buy['B.Date'] = pd.to_datetime(df_residual_buy['B.Date'], format='%d-%m-%Y')

    #This function calculates Net_Amt
    def Amt(quant, unit):
        Net_Amt = quant*unit
        return Net_Amt



    #This function Calculates Profit and Loss
    def pnl(sp, cp):
        a = sp-cp
        return a



    #This function calculates Percentage Profit and Loss
    def per_pnl(sp, cp):
        a = (sp-cp)/cp*100
        return a



    #Adding Net_amt, p_l and per_p_l
    Net_Amt_Residual_buy = Amt(df_residual_buy['B.Quant'],df_residual_buy['B.unitprice'])
    df_residual_buy['B.NetAmt'] = Net_Amt_Residual_buy

    Net_Amt_Residual_Sell = Amt(df_residual_sell['S.Quant'],df_residual_sell['S.unitprice'])
    df_residual_sell['S.NetAmt'] = Net_Amt_Residual_Sell

    Net_Amt_final_buy = Amt(df_q_buy_final['Quant'],df_q_buy_final['unitprice'])
    df_q_buy_final['NetAmt'] = Net_Amt_final_buy

    pl = pnl(df_residual_sell['S.NetAmt'], df_residual_buy['B.NetAmt'])
    df_residual_sell['S.pl'] = pl

    per_pl = per_pnl(df_residual_sell['S.NetAmt'], df_residual_buy['B.NetAmt'])
    df_residual_sell['S.per_pl'] = per_pl

    #sorting
    df_q_buy_final.sort_values(by = 'Date', inplace = True, ascending=True)


    #Merging and sorting
    merged_q_buy_sell = pd.concat([df_residual_buy, df_residual_sell], axis = 1)
    merged_q_buy_sell.sort_values(by = 'S.Date', inplace = True, ascending=True)

    return df_q_buy_final, merged_q_buy_sell





#Function extracts closing price on a given date by user for the valuation purpose
def extracting_stock_price_on_particular_date(sheet, ticker_symbol, date_of_interest):
    n = len(date_of_interest)

    date_of_interest_dt = []
    start_date = []
    end_date = []
    start_date_str = []
    end_date_str = []

    for i in range(n):
        date_of_interest_dt.append(datetime.strptime(date_of_interest[i], '%d-%m-%Y'))
        start_date.append(date_of_interest_dt[i] - timedelta(days=1))
        end_date.append(date_of_interest_dt[i] + timedelta(days=1))

    # Convert the dates to string format
    for i in range(n):
        start_date_str.append(start_date[i].strftime('%Y-%m-%d'))
        end_date_str.append(end_date[i].strftime('%Y-%m-%d'))

    ticker_symbol = [symbol + '.NS' for symbol in ticker_symbol]

    stock_data = []
    stock_data_on_date = []
    for i in range(n):
        stock_data.append(yf.download(ticker_symbol[i], start=start_date_str[i], end=end_date_str[i]))
        stock_data_on_date.append(stock_data[i].loc[stock_data[i].index == date_of_interest_dt[i].strftime('%Y-%m-%d')])

    # Extract the closing price on the specific date
    closing_price = [float(stock_data_on_date[i]['Close'].iloc[0]) for i in range(n)]

    # Create DataFrame for closing prices
    q_closing_price = pd.DataFrame(closing_price, columns=['closing_price'])

    return q_closing_price




#This function combines Valuation, P&l and percentage P&L of a portfolio
def process_and_update_sheet(Amt, Quantity, q_closing_Price):
    
    # Calculate valuation
    val = Quantity['Quant'] * q_closing_Price['closing_price']
    val_df = pd.DataFrame(val, columns=['val'])
    
    # Calculate profit or loss
    p_l = val_df['val'] - Amt['NetAmt']
    p_l_df = pd.DataFrame(p_l, columns=['p_l'])
    
    # Calculate percentage profit or loss
    per_p_l = (p_l_df['p_l'] / Amt['NetAmt']) * 100
    per_p_l_df = pd.DataFrame(per_p_l, columns=['per_p_l'])
    
    # Combine results
    q_final = pd.concat([val_df, p_l_df, per_p_l_df], axis=1)
    return q_final




#This function calculates cagr for the portfolio
def port_cagr(q1):
    n = len(q1['B_Date'])
    d = []
    for i in range(n):
        b = str(q1.loc[i, 'B_Date'])
        c = str(q1.loc[i, 'S_Date'])
        date_format = "%d-%m-%Y"
        sd = datetime.strptime(b, date_format)
        ed = datetime.strptime(c, date_format)
        difference = (ed-sd).days
        d.append(difference)
    d = pd.DataFrame(d, columns=['Days'])
    df = ((q1['Amt_rec']/q1['Amt_payed'])**(365/d['Days']) - 1)*100
    cagr = pd.DataFrame(df, columns=['cagr'])
    return cagr





#This function calculates XIRR value for the portfolio
def calculate_xirr(df, buy_date_col = 'B_Date', sell_date_col = 'S_Date', amt_payed_col = 'Amt_payed', amt_rec_col = 'Amt_rec'):

    df[buy_date_col] = pd.to_datetime(df[buy_date_col], errors='coerce', dayfirst=True)
    df[sell_date_col] = pd.to_datetime(df[sell_date_col], errors='coerce', dayfirst=True)
    
    df[amt_payed_col] = df[amt_payed_col] * (-1)
    
    df_buy = df[[buy_date_col, amt_payed_col]].dropna(subset=[buy_date_col])
    df_sell = df[[sell_date_col, amt_rec_col]].dropna(subset=[sell_date_col])
    

    df_sell.columns = [buy_date_col, amt_payed_col]
    

    df_combined = pd.concat([df_buy, df_sell], axis=0).reset_index(drop=True)
    
    xirr_val = pyxirr.xirr(df_combined[buy_date_col], df_combined[amt_payed_col])
    
    return xirr_val*100




#Calculate Number of days
def No_days(q1):
    q1['B_Date'] = pd.to_datetime(q1['B_Date'], errors='coerce', dayfirst=True)
    q1['S_Date'] = pd.to_datetime(q1['S_Date'], errors='coerce', dayfirst=True)

    days = (q1['S_Date'] - q1['B_Date']).dt.days
    days_df = pd.DataFrame(days, columns=['Days'])
    return days_df




#Function to delete pre written things in Excel sheet
def clear_columns(sheet, col_start, col_end, min_row = 4):
    for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row, min_col=col_start, max_col=col_end):
        for cell in row:
            cell.value = None 




# Writing to the excel file 
def populate_sheet(wb, sheet, q_list, df, col_mapping, file_path):
    def format_and_write_data(mapping, data_frame, columns):
        for i in q_list:
            if i in mapping:
                col = mapping[i]
                for row_num, value in enumerate(df[i], start=4):
                    if i == 'Date' or i == 'S.Date' or i == 'B.Date' and isinstance(value, dt.date):
                        value = value.strftime('%d-%m-%Y')
                    sheet.cell(row=row_num, column=col, value=value)
    
    format_and_write_data(col_mapping, df, q_list)

    wb.save(file_path)


if __name__ == '__main__':


    #reading for buy and sell entry
    wb1, sheet2 = load_worksheet('Buy-sell.xlsx', 'Buy-Sell Entry')
    column_mapping_read1 = {2: 'ISIN', 3: 'symbols', 4: 'BS', 5: 'Quant', 9: 'Date', 13: 'Net_Amt', 14: 'Unit_Price'}
    skip_columns = [6, 7, 8, 10, 11, 12]
    n1 = int(input("Enter last row for your dataset till the point you entered your stocks: \n"))
    q = extract_columns_to_dataframe(sheet2, column_mapping_read1, 3, n1, skip_columns)

    q_buy, q_sell = buy_and_sell_q(q)

    #reading sheet 3
    wb2, sheet3 = load_worksheet('Holdings and valuation.xlsx', 'Holdings, Valuations and P&L ')

    #clearing and writing for buy side entry
    clear_columns(sheet3, 3, 8)
    q_List = ['ISIN','symbols','Quant', 'Date', 'Net_Amt', 'Unit_Price']
    col_mapping1 = {'ISIN': 3, 'symbols': 4, 'Quant': 5, 'Date': 6, 'Net_Amt': 7, 'Unit_Price': 8}
    populate_sheet(wb2, sheet3,q_List,  q_buy, col_mapping1, "Holdings and valuation.xlsx")

    #clearing and Writing for sell side entry
    clear_columns(sheet3, 10, 15)
    q_List = ['ISIN','symbols','Quant', 'Date', 'Net_Amt', 'Unit_Price']
    col_mapping2 = {'ISIN': 10, 'symbols': 11, 'Quant': 12, 'Date': 13, 'Net_Amt': 14, 'Unit_Price': 15}
    populate_sheet(wb2, sheet3,q_List,  q_sell, col_mapping2, "Holdings and valuation.xlsx")

    

    #Wrting it for Portfolio section
    q_buy, q_sell = process_trade_data(q)
    q_buy_formatted, q_sell_formatted = formating_queues(q_buy, q_sell)
    residual_sell, residual_buy, q_buy_final_2D = process_sell_buy_orders(q_sell_formatted, q_buy_formatted)
    df_q_buy_final, merged_q_buy_sell = process_queues(residual_buy, residual_sell, q_buy_final_2D)

    clear_columns(sheet3, 17, 22)
    q_List = ['ISIN','symbol', 'Quant', 'Date', 'unitprice', 'NetAmt']
    col_mapping3 = {'ISIN': 17, 'symbol': 18, 'Quant': 19, 'Date': 20, 'unitprice': 21,'NetAmt': 22}
    populate_sheet(wb2, sheet3, q_List, df_q_buy_final,  col_mapping3, "Holdings and valuation.xlsx")

    #writing for P and L section
    clear_columns(sheet3, 34, 44)
    q_List = ['B.ISIN','B.symbol', 'B.Quant', 'B.Date', 'B.unitprice', 'B.NetAmt', 'B.pl', 'B.per_pl','S.ISIN','S.symbol', 'S.Quant', 'S.Date', 'S.unitprice', 'S.NetAmt', 'S.pl', 'S.per_pl']
    col_mapping4 = {'S.ISIN': 34, 'S.Quant': 36, 'B.Date': 37, 'B.unitprice': 38, 'B.NetAmt': 39, 'S.Date': 40, 'S.unitprice': 41, 'S.NetAmt': 42, 'S.pl': 43, 'S.per_pl': 44, 'S.symbol': 35}
    populate_sheet(wb2, sheet3, q_List, merged_q_buy_sell,  col_mapping4, "Holdings and valuation.xlsx")


    #Reading and writing for Valuation dates
    column_mapping_read2 = {16: 'Valuation_DD', 17: 'Valuation_MM', 18: 'Valuation_YYYY', 19: 'Valuation_Date'}
    n2 = int(input("Enter the last row for the valuation part of buy-sell sheet whose difference should be same as the row difference of your portfolio holdings part of holdings and valuation sheet: \n"))
    q2 = extract_columns_to_dataframe(sheet2, column_mapping_read2, 3, n2, skip_columns)
    q2['Valuation_Date'] = q2.apply(lambda x: f"{int(x['Valuation_DD']):02d}-{int(x['Valuation_MM']):02d}-{int(x['Valuation_YYYY'])}", axis=1)

    clear_columns(sheet3, 23, 26)
    q_List = ['Valuation_DD', 'Valuation_MM', 'Valuation_YYYY', 'Valuation_Date']
    col_mapping5 = {'Valuation_DD':23, 'Valuation_MM':24, 'Valuation_YYYY':25, 'Valuation_Date':26}
    populate_sheet(wb2, sheet3,q_List,  q2, col_mapping5, "Holdings and valuation.xlsx")
    
    #Reading and writing for valuation prices on given Date
    column_mapping_read3 = {18: 'ticker_symbol'}
    n3 = int(input("Enter the last row of your holding and valuation sheet's portfolio holding section: \n"))
    q = extract_columns_to_dataframe(sheet3, column_mapping_read3, 4, n3, skip_columns)
    ticker_symbol = q.values.tolist()
    flattened_ticker_symbol = [item for sublist in ticker_symbol for item in sublist]
    column_mapping_read3 = {26: 'date_of_interest_dt'}
    q = extract_columns_to_dataframe(sheet3, column_mapping_read3, 4, n3, skip_columns)
    date_of_interest_dt = q.values.tolist()
    flattened_date_of_interest_dt = [item for sublist in date_of_interest_dt for item in sublist]

    q_closing_price = extracting_stock_price_on_particular_date(sheet3, flattened_ticker_symbol, flattened_date_of_interest_dt)
    clear_columns(sheet3, 27,27)
    q_List = ['closing_price']
    col_mapping6 = {'closing_price':27}
    populate_sheet(wb2, sheet3,q_List,  q_closing_price, col_mapping6, "Holdings and valuation.xlsx")

    #Quantity and Amount Invested
    column_mapping_read4 = {19: 'Quant'}
    Quantity = extract_columns_to_dataframe(sheet3, column_mapping_read4, 4, n3, skip_columns)

    column_mapping_read5 = {22: 'NetAmt'}
    Amt = extract_columns_to_dataframe(sheet3, column_mapping_read5, 4, n3, skip_columns)

    q_final = process_and_update_sheet(Amt, Quantity, q_closing_price)
    clear_columns(sheet3, 28, 30)
    q_List = ['val', 'p_l', 'per_p_l'] 
    col_mapping7 = {'val':28, 'p_l':29, 'per_p_l':30}
    populate_sheet(wb2, sheet3, q_List, q_final,  col_mapping7, "Holdings and valuation.xlsx")


    #CAGR for portfolio calculation and writing
    column_mapping_read6 = {20: 'B_Date', 22: 'Amt_payed', 26: 'S_Date', 28: 'Amt_rec'}
    skip_columns = [21, 23, 24, 25, 27]
    q = extract_columns_to_dataframe(sheet3, column_mapping_read6, 4, n3, skip_columns)
    cagr = port_cagr(q)

    clear_columns(sheet3, 31, 31)
    q_List = ['cagr']
    col_mapping8 = {'cagr':31}
    populate_sheet(wb2, sheet3,q_List,  cagr, col_mapping8, "Holdings and valuation.xlsx")

    #calculate and write XIRR
    xirr = calculate_xirr(q)
    df = pd.DataFrame([xirr], columns= ["xirr"])
    clear_columns(sheet3, 32, 32)
    q_List = ['xirr']
    col_mapping9 = {'xirr':32}
    populate_sheet(wb2, sheet3,q_List,  df, col_mapping9, "Holdings and valuation.xlsx")

    #calculate XIRR  for P and L
    column_mapping_read7 = {37: 'B_Date', 39: 'Amt_payed', 40: 'S_Date', 42: 'Amt_rec'}
    skip_columns = [38,41]
    n4 = int(input("The last row value of your realized profit and loss part in holding sheet"))
    q = extract_columns_to_dataframe(sheet3, column_mapping_read7, 4, n4, skip_columns)
    xirr_pl = calculate_xirr(q)

    #write XIRR P and L in excel
    df = pd.DataFrame([xirr_pl], columns= ["xirr_pl"])
    clear_columns(sheet3, 45, 45)
    q_List = ['xirr_pl']
    col_mapping10 = {'xirr_pl':45}
    populate_sheet(wb2, sheet3,q_List,  df, col_mapping10, "Holdings and valuation.xlsx")

    #calculate Number of days(differene between buy date and sell date)
    Days = No_days(q)
    LTCG = []
    STCG = []


    #Extracting for p and l column
    column_mapping_read8 = {43: 'Profit/Loss'}
    q = extract_columns_to_dataframe(sheet3, column_mapping_read8, 4, n4, skip_columns)

    result = pd.concat([Days, q], axis = 1)
    n = len(Days)

    #Separating stcg and ltcg
    for i in range(n):
        if result.loc[i, 'Days'] > 365:
            LTCG.append(result.loc[i, 'Profit/Loss'])
            STCG.append(0)
        else:
            LTCG.append(0)
            STCG.append(result.loc[i, 'Profit/Loss'])

    

    #Making dataframe
    LTCG_STCG = pd.DataFrame(columns=['LTCG', 'STCG'])
    for i in range(n):
        LTCG_STCG.loc[i] = [LTCG[i], STCG[i]]
    LTCG_STCG_Days = pd.concat([Days, LTCG_STCG], axis = 1)

    clear_columns(sheet3, 48, 50)
    q_List = ['Days', 'LTCG', 'STCG']
    col_mapping11 = {'Days':48, 'LTCG':49, 'STCG':50}
    populate_sheet(wb2, sheet3,q_List,  LTCG_STCG_Days, col_mapping11, "Holdings and valuation.xlsx")