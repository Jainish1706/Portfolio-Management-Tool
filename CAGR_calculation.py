#Importing libraries
import pandas as pd
import openpyxl
from datetime import datetime



def load_sheet_data(file_path, sheet_name, min_row, max_row):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        data = sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True)
        return wb, sheet, data
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return None



def create_dataframe_from_sheet(sheet, col_mapping, max_row, min_row=4):
    data = {name: [] for name in col_mapping.values()}

    # Loop through specified columns
    for col_index, col_name in col_mapping.items():
        for row in sheet.iter_rows(min_col=col_index, max_col=col_index, min_row=min_row, max_row=max_row, values_only=True):
            for cell in row:
                data[col_name].append(cell)

    # Create DataFrame
    df = pd.DataFrame(data)
    return df




def port_cagr(q1):
    n = len(q1['B_Date'])
    d = []
    for i in range(n):
        b = str(q1.loc[i, 'B_Date'])
        c = str(q1.loc[i, 'S_Date'])
        date_format = "%d-%m-%Y"
        sd = datetime.strptime(b, date_format)
        ed = datetime.strptime(c, date_format)
        difference = (ed-sd).days
        d.append(difference)
    d = pd.DataFrame(d, columns=['Days'])
    df = ((q1['Amt_rec']/q1['Amt_payed'])**(365/d['Days']) - 1)*100
    cagr = pd.DataFrame(df, columns=['cagr'])
    return cagr





def clear_columns(sheet, col_start1, col_end1, col_start2, col_end2, min_row = 4):
    for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row, min_col=col_start1, max_col=col_end1):
        for cell in row:
            cell.value = None 



def write_column_to_sheet(wb, sheet, data, start_row, column, file_path):
    for row_num, value in enumerate(data, start=start_row):
        sheet.cell(row=row_num, column=column, value=value)
    
    wb.save(file_path)