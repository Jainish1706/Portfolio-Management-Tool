#Importing Libraries
import pandas as pd
import openpyxl
import pyxirr



def load_sheet_data(file_path, sheet_name, min_row, max_row):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        data = sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True)
        return wb, sheet, data
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
    


def create_dataframe_from_sheet(sheet, col_mapping, max_row, min_row=4):
    data = {name: [] for name in col_mapping.values()}

    # Loop through specified columns
    for col_index, col_name in col_mapping.items():
        for row in sheet.iter_rows(min_col=col_index, max_col=col_index, min_row=min_row, max_row=max_row, values_only=True):
            for cell in row:
                data[col_name].append(cell)

    # Create DataFrame
    df = pd.DataFrame(data)
    return df





def process_transactions(df, buy_date_col='B_Date', sell_date_col='S_Date', amt_payed_col='Amt_payed', amt_rec_col='Amt_rec'):

    df[buy_date_col] = pd.to_datetime(df[buy_date_col], errors='coerce', dayfirst=True)
    df[sell_date_col] = pd.to_datetime(df[sell_date_col], errors='coerce', dayfirst=True)


    df[amt_payed_col] = df[amt_payed_col] * (-1)


    q1 = df[[buy_date_col, amt_payed_col]].dropna(subset=[buy_date_col])
    q2 = df[[sell_date_col, amt_rec_col]].dropna(subset=[sell_date_col])


    q2.columns = [buy_date_col, amt_payed_col]


    combined_df = pd.concat([q1, q2], axis=0).reset_index(drop=True)

    return combined_df





def update_xirr_in_workbook(df, workbook_path, sheet_name, date_col, amount_col, cell_location):

    xirr_val = pyxirr.xirr(df[date_col], df[amount_col])
    
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb[sheet_name]
    
    sheet[cell_location] = xirr_val * 100
    
    wb.save(workbook_path)