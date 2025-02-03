# Importing Libraries
import pandas as pd
import openpyxl
import yfinance as yf
from datetime import datetime, timedelta

def load_sheet_data(file_path, sheet_name, min_row, max_row):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        data = sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True)
        return wb, sheet, data
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return None



def extract_valuation_date(sheet, max_row):

    Valuation_DD = []
    Valuation_MM = []
    Valuation_YYYY = []
    Valuation_Date = []

    for i in range(16, 20):
        for row in sheet.iter_rows(min_col = i, max_col = i, min_row = 3, values_only = True, max_row=max_row):
            for cell in row:
                if i == 16:
                    Valuation_DD.append(cell)
                elif i == 17:
                    Valuation_MM.append(cell)
                elif i == 18:
                    Valuation_YYYY.append(cell)
                elif i == 19:
                    Valuation_Date.append(cell)
    
    df = pd.DataFrame(columns=['Valuation_DD', 'Valuation_MM', 'Valuation_YYYY'])
    df['Valuation_DD'] = Valuation_DD
    df['Valuation_MM'] = Valuation_MM
    df['Valuation_YYYY'] = Valuation_YYYY

    # Combine the date parts into the required format
    df['Valuation_Date'] = df.apply(lambda x: f"{int(x['Valuation_DD']):02d}-{int(x['Valuation_MM']):02d}-{int(x['Valuation_YYYY'])}", axis=1)

    return df



def clear_columns(sheet, col_start1, col_end1, col_start2, col_end2, min_row = 4):
    for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row, min_col=col_start1, max_col=col_end1):
        for cell in row:
            cell.value = None 



def write_fomatted_dates_to_excel(wb, file_path, sheet, df, q_list):
    excel_columns = {'Valuation_DD':23, 'Valuation_MM':24, 'Valuation_YYYY':25, 'Valuation_Date':26}
    for col_name in q_list:
        if col_name in df.columns:
            for row_num, value in enumerate(df[col_name], start = 4):
                sheet.cell(row=row_num, column=excel_columns[col_name], value=value)

    wb.save(file_path)




def extracting_stock_price_on_particular_date(sheet, max_row):
    ticker_symbol = []
    date_of_interest = []
    Amt_invested = []
    Quant = []

    for i in range(18, 27):
        if 20 <= i <= 21 or 23 <= i <= 25:  # Skip columns 20-21 and 23-25
            continue
        else:
            for row in sheet.iter_rows(min_col=i, max_col=i, min_row=4, values_only=True, max_row=9):
                for cell in row:
                    if i == 18:
                        ticker_symbol.append(cell)
                    elif i == 26:
                        date_of_interest.append(cell)
                    elif i == 22:
                        Amt_invested.append(cell)
                    elif i == 19:
                        Quant.append(cell)
        
    n = len(date_of_interest)

    date_of_interest_dt = []
    start_date = []
    end_date = []
    start_date_str = []
    end_date_str = []

    for i in range(n):
        date_of_interest_dt.append(datetime.strptime(date_of_interest[i], '%d-%m-%Y'))
        start_date.append(date_of_interest_dt[i] - timedelta(days=1))
        end_date.append(date_of_interest_dt[i] + timedelta(days=1))

    # Convert the dates to string format
    for i in range(n):
        start_date_str.append(start_date[i].strftime('%Y-%m-%d'))
        end_date_str.append(end_date[i].strftime('%Y-%m-%d'))

    ticker_symbol = [symbol + '.NS' for symbol in ticker_symbol]

    stock_data = []
    stock_data_on_date = []
    for i in range(n):
        stock_data.append(yf.download(ticker_symbol[i], start=start_date_str[i], end=end_date_str[i]))
        stock_data_on_date.append(stock_data[i].loc[stock_data[i].index == date_of_interest_dt[i].strftime('%Y-%m-%d')])

    # Extract the closing price on the specific date
    closing_price = [float(stock_data_on_date[i]['Close'].iloc[0]) for i in range(n)]

    # Create DataFrame for closing prices
    q_closing_price = pd.DataFrame(closing_price, columns=['closing_price'])

    return q_closing_price




def update_sheet_with_values(sheet, q_list, q_closing_Price, start_row=4):
    
    column_mapping = {'closing_price': 27}                 

    for key in q_list:
        if key in column_mapping:
            column = column_mapping[key]
            for row_num, value in enumerate(q_closing_Price[key], start=start_row):
                sheet.cell(row=row_num, column=column, value=value)




def process_and_update_sheet(Amt_invested, Quant, q_closing_Price, sheet, wb, q_list1, column_map):
    
    # Create DataFrames
    Amt = pd.DataFrame(Amt_invested, columns=['NetAmt'])
    Quantity = pd.DataFrame(Quant, columns=['Quant'])
    
    # Calculate valuation
    val = Quantity['Quant'] * q_closing_Price['closing_price']
    val_df = pd.DataFrame(val, columns=['val'])
    
    # Calculate profit or loss
    p_l = val_df['val'] - Amt['NetAmt']
    p_l_df = pd.DataFrame(p_l, columns=['p_l'])
    
    # Calculate percentage profit or loss
    per_p_l = (p_l_df['p_l'] / Amt['NetAmt']) * 100
    per_p_l_df = pd.DataFrame(per_p_l, columns=['per_p_l'])
    
    # Combine results
    q_final = pd.concat([val_df, p_l_df, per_p_l_df], axis=1)
    
    # Update Excel sheet
    for column_name in q_list1:
        if column_name in q_final.columns:
            col_number = column_map.get(column_name)
            if col_number is not None:
                for row_num, value in enumerate(q_final[column_name], start=4):
                    sheet.cell(row=row_num, column=col_number, value=value)
    
    # Save the workbook
    wb.save("Holdings and valuation.xlsx")